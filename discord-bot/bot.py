import discord
from discord.ext import commands, tasks
from flask import Flask, request, jsonify
import sqlalchemy
from sqlalchemy import create_engine, text, Column, String, Integer, Float, DateTime, Boolean, JSON
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from datetime import datetime, timedelta, timezone
import hashlib
import random
import asyncio
import os
from dotenv import load_dotenv
import requests
import json
import time
from typing import Optional
from discord.ui import View, Button, Modal, TextInput
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

# 載入環境變數
load_dotenv()

# Discord Bot 設定
intents = discord.Intents.default()
intents.message_content = True
intents.guilds = True
intents.voice_channels = True
intents.members = True

bot = commands.Bot(command_prefix='!', intents=intents)

# 從環境變數獲取 Discord Token
DISCORD_TOKEN = os.getenv('DISCORD_TOKEN')
GUILD_ID = int(os.getenv('DISCORD_GUILD_ID', '0'))
ADMIN_CHANNEL_ID = int(os.getenv('ADMIN_CHANNEL_ID', '0'))
ADMIN_USER_ID = int(os.getenv('ADMIN_USER_ID', '0'))

# 檢查必要的環境變數
if not DISCORD_TOKEN:
    print("❌ 錯誤：未找到 DISCORD_TOKEN 環境變數")
    exit(1)

if GUILD_ID == 0:
    print("❌ 錯誤：未找到 DISCORD_GUILD_ID 環境變數")
    exit(1)

if ADMIN_CHANNEL_ID == 0:
    print("❌ 錯誤：未找到 ADMIN_CHANNEL_ID 環境變數")
    exit(1)

if ADMIN_USER_ID == 0:
    print("❌ 錯誤：未找到 ADMIN_USER_ID 環境變數")
    exit(1)

print(f"✅ Discord Token: {DISCORD_TOKEN[:10]}...")
print(f"✅ Guild ID: {GUILD_ID}")
print(f"✅ Admin Channel ID: {ADMIN_CHANNEL_ID}")
print(f"✅ Admin User ID: {ADMIN_USER_ID}")

# 資料庫設定
DATABASE_URL = os.getenv('DATABASE_URL')
if not DATABASE_URL:
    print("❌ 錯誤：未找到 DATABASE_URL 環境變數")
    exit(1)

try:
    engine = create_engine(DATABASE_URL)
    print("✅ 資料庫連線成功")
except Exception as e:
    print(f"❌ 資料庫連線失敗: {e}")
    exit(1)

# 創建 Session 類
Session = sessionmaker(bind=engine)

# 可愛物品列表
CUTE_ITEMS = [
    "蝴蝶結", "小狗", "小貓", "小熊", "小兔", "小鳥", "小魚", "小花", 
    "小樹", "小星", "小月", "小太陽", "小雲", "小彩虹", "小愛心",
    "小鑽石", "小皇冠", "小翅膀", "小鈴鐺", "小糖果", "小蛋糕",
    "小冰淇淋", "小氣球", "小禮物", "小寶石", "小珍珠", "小貝殼"
]

# 檢查間隔（秒）
CHECK_INTERVAL = 30

# 評價系統全局變數
evaluated_records = set()
pending_ratings = {}
rating_submitted_users = {}  # 追蹤每個記錄的已提交評價用戶 {record_id: set(user_ids)}
rating_text_channels = {}  # 追蹤每個記錄的文字頻道 {record_id: text_channel}
rating_channel_created_time = {}  # 追蹤每個記錄的文字頻道創建時間 {record_id: timestamp}
rating_notification_cache = {}  # 緩存評價通知 {record_id: {'ratings': [rating_data], 'user1_id': str, 'user2_id': str, 'timer': task}}
pairing_record_sent = set()  # 追蹤已發送配對紀錄的 record_id，避免重複發送

# 創建 Discord 頻道的函數
def create_booking_text_channel(guild, booking_id, customer_name, partner_name, is_instant_booking=False):
    """創建預約文字頻道"""
    try:
        # 使用 MD5 雜湊確保一致性
        hash_obj = hashlib.md5(booking_id.encode())
        hash_hex = hash_obj.hexdigest()
        cute_item = CUTE_ITEMS[int(hash_hex[:2], 16) % len(CUTE_ITEMS)]
        
        if is_instant_booking:
            channel_name = f"🔥{cute_item}-{customer_name}-{partner_name}"
            else:
            channel_name = f"📝{cute_item}-{customer_name}-{partner_name}"
        
        # 檢查頻道是否已存在
        existing_channel = discord.utils.get(guild.text_channels, name=channel_name)
        if existing_channel:
            print(f"⚠️ 文字頻道已存在: {channel_name}")
            return existing_channel
        
        # 創建頻道
        channel = guild.create_text_channel(
            channel_name,
            category=None,  # 不指定分類
            topic=f"預約頻道 - 客戶: {customer_name}, 夥伴: {partner_name}"
        )
        print(f"✅ 創建文字頻道: {channel_name}")
        return channel
        
    except Exception as e:
        print(f"❌ 創建文字頻道失敗: {e}")
        return None

def create_booking_voice_channel(guild, booking_id, customer_name, partner_name):
    """創建預約語音頻道"""
    try:
        # 使用相同的 MD5 雜湊確保一致性
        hash_obj = hashlib.md5(booking_id.encode())
        hash_hex = hash_obj.hexdigest()
        cute_item = CUTE_ITEMS[int(hash_hex[:2], 16) % len(CUTE_ITEMS)]
        
        channel_name = f"🎤{cute_item}-{customer_name}-{partner_name}"
        
        # 檢查頻道是否已存在
        existing_channel = discord.utils.get(guild.voice_channels, name=channel_name)
        if existing_channel:
            print(f"⚠️ 語音頻道已存在: {channel_name}")
            return existing_channel
        
        # 創建頻道
        channel = guild.create_voice_channel(
            channel_name,
            category=None,  # 不指定分類
            bitrate=64000  # 設置音質
        )
        print(f"✅ 創建語音頻道: {channel_name}")
        return channel
        
    except Exception as e:
        print(f"❌ 創建語音頻道失敗: {e}")
        return None

# 檢查資料庫連線健康的函數
@tasks.loop(minutes=5)
async def database_health_check():
    """定期檢查資料庫連線健康狀態"""
    try:
        session = Session()
        session.execute(text("SELECT 1"))
        session.close()
        print("✅ 資料庫連線正常")
    except Exception as e:
        print(f"❌ 資料庫連線異常: {e}")
        # 嘗試重新初始化引擎
        try:
            global engine
            engine.dispose()
            engine = create_engine(DATABASE_URL)
            print("✅ 資料庫引擎重新初始化成功")
        except Exception as e2:
            print(f"❌ 資料庫引擎重新初始化失敗: {e2}")

# 計算推薦獎勵的函數
async def calculate_referral_earnings(booking_id):
    """計算推薦獎勵"""
    try:
        response = requests.post('https://peiplay.vercel.app/api/partners/referral/calculate-earnings', 
                               json={'bookingId': booking_id})
        if response.status_code == 200:
            print(f"✅ 推薦獎勵計算成功: {booking_id}")
                else:
            print(f"⚠️ 推薦獎勵計算失敗: {booking_id}, 狀態碼: {response.status_code}")
    except Exception as e:
        print(f"❌ 推薦獎勵計算錯誤: {e}")

# 檢查待審核項目的函數
@tasks.loop(hours=6)
async def check_pending_reviews():
    """檢查待審核的夥伴申請和提領申請"""
    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            print("❌ 找不到 Discord 伺服器")
            return
        
        admin_channel = guild.get_channel(ADMIN_CHANNEL_ID)
        if not admin_channel:
            print("❌ 找不到管理員頻道")
            return
        
        session = Session()
        
        # 檢查待審核的夥伴申請
        pending_partners = session.execute(text("""
            SELECT COUNT(*) as count FROM "Partner" 
            WHERE status = 'PENDING'
        """)).fetchone()
        
        # 檢查待審核的提領申請
        pending_withdrawals = session.execute(text("""
            SELECT COUNT(*) as count FROM "WithdrawalRequest" 
            WHERE status = 'PENDING'
        """)).fetchone()
        
        session.close()
        
        # 如果有待審核項目，發送通知
        if pending_partners.count > 0 or pending_withdrawals.count > 0:
            message = "🔔 **管理員通知**\n\n"
            if pending_partners.count > 0:
                message += f"📋 待審核夥伴申請: {pending_partners.count} 件\n"
            if pending_withdrawals.count > 0:
                message += f"💰 待審核提領申請: {pending_withdrawals.count} 件\n"
            message += "\n請及時處理待審核項目。"
            
            await admin_channel.send(message)
            print(f"✅ 發送待審核通知: 夥伴 {pending_partners.count} 件, 提領 {pending_withdrawals.count} 件")
        
    except Exception as e:
        print(f"❌ 檢查待審核項目時發生錯誤: {e}")

# 新的預約流程檢查函數
async def check_early_communication_channels(guild, now):
    """檢查需要創建提前溝通文字頻道的預約（預約確認後）"""
    try:
        session = Session()
        
        # 查找已確認但還沒有提前溝通頻道的預約
        bookings = session.execute(text("""
            SELECT b.id, b.customerId, b.scheduleId, b.discordEarlyTextChannelId,
                   c.name as customer_name, p.name as partner_name, b."paymentInfo"
            FROM "Booking" b
            JOIN "Customer" c ON b.customerId = c.id
            JOIN "Schedule" s ON b.scheduleId = s.id
            JOIN "Partner" p ON s.partnerId = p.id
            WHERE b.status = 'CONFIRMED' 
            AND b.discordEarlyTextChannelId IS NULL
            AND b.createdAt <= :now
        """), {'now': now}).fetchall()
        
        session.close()
        
        for booking in bookings:
            try:
                # 判斷是否為即時預約
                is_instant_booking = False
                if booking.paymentInfo and isinstance(booking.paymentInfo, dict):
                    is_instant_booking = booking.paymentInfo.get('isInstantBooking') == 'true'
                
                # 創建提前溝通文字頻道
                channel = create_booking_text_channel(
                    guild, 
                    booking.id, 
                    booking.customer_name, 
                    booking.partner_name,
                    is_instant_booking
                )
                
                if channel:
                    # 更新資料庫
                    session = Session()
                    session.execute(text("""
                        UPDATE "Booking" 
                        SET "discordEarlyTextChannelId" = :channel_id
                        WHERE id = :booking_id
                    """), {'channel_id': str(channel.id), 'booking_id': booking.id})
                    session.commit()
                    session.close()
                    
                    # 發送歡迎訊息
                            embed = discord.Embed(
                        title="🎮 預約確認",
                        description=f"嗨 {booking.customer_name}！你的預約已確認，夥伴 {booking.partner_name} 將在預約時間與你聯繫。",
                        color=0x00ff00
                    )
                    embed.add_field(name="📅 預約時間", value="請等待夥伴確認具體時間", inline=False)
                    embed.add_field(name="💬 溝通方式", value="此頻道用於預約前的溝通", inline=False)
                    
                    await channel.send(embed=embed)
                    print(f"✅ 創建提前溝通頻道: {booking.id}")
                
                except Exception as e:
                print(f"❌ 處理預約 {booking.id} 時發生錯誤: {e}")
        
    except Exception as e:
        print(f"❌ 檢查提前溝通頻道時發生錯誤: {e}")

async def check_voice_channel_creation(guild, now):
    """檢查需要創建語音頻道的預約（開始前5分鐘）"""
    try:
        session = Session()
        
        # 查找需要創建語音頻道的預約（開始前5分鐘，且有提前溝通頻道但沒有語音頻道）
        five_minutes_later = now + timedelta(minutes=5)
        bookings = session.execute(text("""
            SELECT b.id, b.customerId, b.scheduleId, b.discordEarlyTextChannelId, b.discordTextChannelId, b.discordVoiceChannelId,
                   c.name as customer_name, p.name as partner_name, s.startTime, b."paymentInfo"
            FROM "Booking" b
            JOIN "Customer" c ON b.customerId = c.id
            JOIN "Schedule" s ON b.scheduleId = s.id
            JOIN "Partner" p ON s.partnerId = p.id
            WHERE b.status = 'CONFIRMED'
            AND b.discordEarlyTextChannelId IS NOT NULL
            AND b.discordVoiceChannelId IS NULL
            AND s.startTime <= :five_minutes_later
            AND s.startTime > :now
        """), {'five_minutes_later': five_minutes_later, 'now': now}).fetchall()
        
        session.close()
        
        for booking in bookings:
            try:
                # 創建語音頻道
                voice_channel = create_booking_voice_channel(
                    guild, 
                    booking.id, 
                    booking.customer_name, 
                    booking.partner_name
                )
                
                if voice_channel:
                    # 創建正式文字頻道
                    text_channel = create_booking_text_channel(
                        guild, 
                        booking.id, 
                        booking.customer_name, 
                        booking.partner_name
                    )
                    
                if text_channel:
                        # 更新資料庫
                        session = Session()
                        session.execute(text("""
                            UPDATE "Booking" 
                            SET "discordVoiceChannelId" = :voice_id, "discordTextChannelId" = :text_id
                            WHERE id = :booking_id
                        """), {
                            'voice_id': str(voice_channel.id), 
                            'text_id': str(text_channel.id),
                            'booking_id': booking.id
                        })
                        session.commit()
                        session.close()
                        
                        # 刪除提前溝通頻道
                        try:
                            early_channel = guild.get_channel(int(booking.discordEarlyTextChannelId))
                            if early_channel:
                                await early_channel.delete()
                                print(f"✅ 刪除提前溝通頻道: {booking.id}")
                        except Exception as e:
                            print(f"⚠️ 刪除提前溝通頻道失敗: {e}")
                        
                        # 在正式文字頻道發送歡迎訊息
                        embed = discord.Embed(
                            title="🎮 預約開始",
                            description=f"預約即將開始！請進入語音頻道開始遊戲。",
                            color=0x0099ff
                        )
                        embed.add_field(name="🎤 語音頻道", value=f"請點擊 {voice_channel.mention} 進入", inline=False)
                        embed.add_field(name="⏰ 開始時間", value=f"<t:{int(booking.startTime.timestamp())}:R>", inline=False)
                        
                        await text_channel.send(embed=embed)
                        print(f"✅ 創建正式頻道: {booking.id}")
                
            except Exception as e:
                print(f"❌ 處理預約 {booking.id} 時發生錯誤: {e}")
        
    except Exception as e:
        print(f"❌ 檢查語音頻道創建時發生錯誤: {e}")

async def check_extension_buttons(guild, now):
    """檢查需要顯示延長按鈕的預約（結束前10分鐘）"""
    try:
        session = Session()
        
        # 查找需要顯示延長按鈕的預約（結束前10分鐘，且還沒有顯示過）
        ten_minutes_later = now + timedelta(minutes=10)
        bookings = session.execute(text("""
            SELECT b.id, b.discordTextChannelId, s.endTime
                FROM "Booking" b
            JOIN "Schedule" s ON b.scheduleId = s.id
                WHERE b.status = 'CONFIRMED'
            AND b.discordTextChannelId IS NOT NULL
            AND b.extensionButtonShown = false
            AND s.endTime <= :ten_minutes_later
            AND s.endTime > :now
        """), {'ten_minutes_later': ten_minutes_later, 'now': now}).fetchall()
        
        session.close()
        
        for booking in bookings:
            try:
                text_channel = guild.get_channel(int(booking.discordTextChannelId))
                if text_channel:
                    # 發送延長按鈕
                    embed = discord.Embed(
                        title="⏰ 預約即將結束",
                        description="預約還有 10 分鐘結束，是否需要延長 5 分鐘？",
                        color=0xff9900
                    )
                    
                    view = discord.ui.View()
                    extend_button = discord.ui.Button(
                        label="延長 5 分鐘",
                        style=discord.ButtonStyle.primary,
                        custom_id=f"extend_booking_{booking.id}"
                    )
                    view.add_item(extend_button)
                    
                    await text_channel.send(embed=embed, view=view)
                    
                    # 更新資料庫
                    session = Session()
                    session.execute(text("""
                        UPDATE "Booking" 
                        SET "extensionButtonShown" = true
                        WHERE id = :booking_id
                    """), {'booking_id': booking.id})
                    session.commit()
                    session.close()
                    
                    print(f"✅ 顯示延長按鈕: {booking.id}")
                        
                except Exception as e:
                print(f"❌ 處理預約 {booking.id} 時發生錯誤: {e}")
                    
    except Exception as e:
        print(f"❌ 檢查延長按鈕時發生錯誤: {e}")

async def check_voice_channel_cleanup(guild, now):
    """檢查需要結束語音頻道的預約（時間結束）"""
    try:
        session = Session()
        
        # 查找需要結束語音頻道的預約
        bookings = session.execute(text("""
            SELECT b.id, b.discordVoiceChannelId, b.discordTextChannelId, b.ratingCompleted,
                   c.name as customer_name, p.name as partner_name, s.endTime
        FROM "Booking" b
            JOIN "Customer" c ON b.customerId = c.id
            JOIN "Schedule" s ON b.scheduleId = s.id
            JOIN "Partner" p ON s.partnerId = p.id
            WHERE b.status = 'CONFIRMED'
            AND b.discordVoiceChannelId IS NOT NULL
            AND s.endTime <= :now
        """), {'now': now}).fetchall()
        
        session.close()
        
        for booking in bookings:
            try:
                # 刪除語音頻道
                voice_channel = guild.get_channel(int(booking.discordVoiceChannelId))
                if voice_channel and not voice_channel.deleted:
                            await voice_channel.delete()
                    print(f"✅ 刪除語音頻道: {booking.id}")
                
                # 在文字頻道顯示評價系統
                text_channel = guild.get_channel(int(booking.discordTextChannelId))
                if text_channel and not booking.ratingCompleted:
                    embed = discord.Embed(
                        title="⭐ 預約結束",
                        description=f"預約已結束，請為夥伴 {booking.partner_name} 評分。\n點選下方按鈕選擇 1-5 顆星評價：",
                        color=0x9932cc
                    )
                    
                    view = discord.ui.View()
                    for i in range(1, 6):
                        # 根據星級數量顯示不同數量的星號
                        stars = "⭐" * i
                        button = discord.ui.Button(
                            label=f"{i} 顆星 {stars}",
                            style=discord.ButtonStyle.secondary,
                            custom_id=f"rate_{booking.id}_{i}"
                        )
                        view.add_item(button)
                    
                    await text_channel.send(embed=embed, view=view)
                    print(f"✅ 顯示評價系統: {booking.id}")
                
            except Exception as e:
                print(f"❌ 處理預約 {booking.id} 時發生錯誤: {e}")
        
    except Exception as e:
        print(f"❌ 檢查語音頻道清理時發生錯誤: {e}")

async def check_text_channel_cleanup(guild, now):
    """檢查需要清理文字頻道的預約（評價完成後）"""
    try:
        session = Session()
        
        # 查找需要清理文字頻道的預約（評價完成且文字頻道未清理）
        bookings = session.execute(text("""
            SELECT b.id, b.discordTextChannelId, b.ratingCompleted, b.textChannelCleaned
        FROM "Booking" b
            WHERE b.ratingCompleted = true
            AND b.textChannelCleaned = false
            AND b.discordTextChannelId IS NOT NULL
        """)).fetchall()
        
        session.close()
        
        for booking in bookings:
            try:
                # 刪除文字頻道
                text_channel = guild.get_channel(int(booking.discordTextChannelId))
                if text_channel and not text_channel.deleted:
                    await text_channel.delete()
                    print(f"✅ 刪除文字頻道: {booking.id}")
                
                # 更新資料庫
                session = Session()
                session.execute(text("""
                    UPDATE "Booking" 
                    SET "textChannelCleaned" = true
                    WHERE id = :booking_id
                """), {'booking_id': booking.id})
                session.commit()
                session.close()
                
            except Exception as e:
                print(f"❌ 處理預約 {booking.id} 時發生錯誤: {e}")
                
    except Exception as e:
        print(f"❌ 檢查文字頻道清理時發生錯誤: {e}")

async def cleanup_expired_group_text_channels(guild):
    """清理過期的群組預約文字頻道（匿名文字區）"""
    try:
        # 查找所有名為"匿名文字區"的文字頻道
        text_channels = [ch for ch in guild.text_channels if ch.name == "🔒匿名文字區" or "匿名文字區" in ch.name]
        
        if not text_channels:
            return
        
        # 通過 API 獲取所有已結束的群組預約
        try:
            # 調用 Next.js API 來獲取已結束的群組預約
            nextjs_url = os.getenv('NEXTAUTH_URL', 'http://localhost:3004')
            try:
                response = requests.get(
                    f"{nextjs_url}/api/cron/group-booking-status",
                    timeout=10
                )
                if response.status_code == 200:
                    result = response.json()
                    print(f"📊 群組預約狀態檢查完成: {result}")
            except Exception as e:
                print(f"⚠️ 無法連接到 Next.js API: {e}")
            
            # 直接檢查頻道的創建時間和最後活動時間
            # 如果頻道創建時間超過 6 小時，且最後一條訊息超過 1 小時，則刪除
            now = datetime.now(timezone.utc)
            
            for channel in text_channels:
                try:
                    # 檢查頻道創建時間
                    channel_age = (now - channel.created_at.replace(tzinfo=timezone.utc)).total_seconds()
                    
                    # 獲取最後一條訊息
                    last_message = None
                    try:
                        async for message in channel.history(limit=1):
                            last_message = message
                            break
                    except:
                        pass
                    
                    # 如果頻道創建超過 6 小時
                    if channel_age > 6 * 60 * 60:  # 6 小時
                        should_delete = False
                        
                        if last_message:
                            # 如果最後一條訊息超過 1 小時，則刪除
                            message_age = (now - last_message.created_at.replace(tzinfo=timezone.utc)).total_seconds()
                            if message_age > 1 * 60 * 60:  # 1 小時
                                should_delete = True
                        else:
                            # 如果沒有訊息，且創建超過 6 小時，則刪除
                            should_delete = True
                        
                        if should_delete:
                            await channel.delete()
                            print(f"✅ 已刪除過期群組文字頻道: {channel.name} ({channel.id})")
                except discord.errors.NotFound:
                    # 頻道已經被刪除，跳過
                    pass
                except Exception as e:
                    print(f"❌ 清理頻道 {channel.id} 時發生錯誤: {e}")
        except Exception as e:
            print(f"❌ 清理群組文字頻道時發生錯誤: {e}")
            
    except Exception as e:
        print(f"❌ 檢查群組文字頻道清理時發生錯誤: {e}")

async def check_expired_rating_channels(guild, now):
    """檢查評價系統文字頻道是否超過5分鐘未完成評價"""
    try:
        expired_channels = []
        current_time = datetime.now(timezone.utc)
        
        for record_id, created_time in list(rating_channel_created_time.items()):
            # 檢查是否超過5分鐘（300秒）
            time_diff = (current_time - created_time).total_seconds()
            if time_diff >= 300:  # 5分鐘
                # 檢查是否還有未完成的評價
                if record_id in rating_text_channels:
                    text_channel = rating_text_channels[record_id]
                    # 檢查是否所有用戶都已提交
                    submitted_users = rating_submitted_users.get(record_id, set())
                    
                    session = Session()
                    try:
                        result = session.execute(text("""
                            SELECT "user1Id", "user2Id" 
                            FROM "PairingRecord" 
                            WHERE id = :record_id
                        """), {"record_id": record_id}).fetchone()
                        
                        if result:
                            user1_id = result[0]
                            user2_id = result[1]
                            
                            user1_submitted = str(user1_id) in submitted_users
                            user2_submitted = str(user2_id) in submitted_users
                            is_single_user = str(user1_id) == str(user2_id)
                            
                            # 如果還有用戶未提交，且超過5分鐘，則刪除頻道
                            if not ((user1_submitted and user2_submitted) or (is_single_user and user1_submitted)):
                                expired_channels.append((record_id, text_channel))
                    finally:
                        session.close()
        
        # 刪除過期的頻道
        for record_id, text_channel in expired_channels:
            try:
                if text_channel and not text_channel.deleted:
                    await text_channel.delete()
                    print(f"✅ 5分鐘內未完成評價，已刪除文字頻道: {text_channel.name} (record_id: {record_id})")
                    # 清理追蹤
                    rating_text_channels.pop(record_id, None)
                    rating_channel_created_time.pop(record_id, None)
                    
                    # 發送配對紀錄到管理員頻道（即使沒有評價）
                    await send_pairing_record_to_admin(record_id)
            except Exception as e:
                print(f"❌ 刪除過期評價頻道失敗: {e}")
                # 即使刪除失敗，也清理追蹤
                rating_text_channels.pop(record_id, None)
                rating_channel_created_time.pop(record_id, None)
                
                # 即使刪除失敗，也發送配對紀錄
                try:
                    await send_pairing_record_to_admin(record_id)
                except:
                    pass
                
    except Exception as e:
        print(f"❌ 檢查過期評價頻道時發生錯誤: {e}")

# 主要的預約檢查任務
@tasks.loop(seconds=CHECK_INTERVAL)
async def check_bookings():
    """定期檢查預約狀態並管理 Discord 頻道"""
    await bot.wait_until_ready()

    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            print("❌ 找不到 Discord 伺服器")
            return
        
        now = datetime.now(timezone.utc)
        
        # 1. 檢查需要創建提前溝通文字頻道的預約（預約確認後）
        await check_early_communication_channels(guild, now)
        
        # 2. 檢查需要創建語音頻道的預約（開始前5分鐘）
        await check_voice_channel_creation(guild, now)
        
        # 3. 檢查需要顯示延長按鈕的預約（結束前10分鐘）
        await check_extension_buttons(guild, now)
        
        # 4. 檢查需要結束語音頻道的預約（時間結束）
        await check_voice_channel_cleanup(guild, now)
        
        # 5. 檢查需要清理文字頻道的預約（評價完成後）
        await check_text_channel_cleanup(guild, now)
        
        # 6. 清理過期的群組預約文字頻道
        await cleanup_expired_group_text_channels(guild)
        
        # 7. 檢查評價系統文字頻道是否超過5分鐘未完成評價
        await check_expired_rating_channels(guild, now)
        
    except Exception as e:
        print(f"❌ 檢查預約時發生錯誤: {e}")

# 檢查新預約的任務
@tasks.loop(seconds=CHECK_INTERVAL)
async def check_new_bookings():
    """檢查新的預約並創建 Discord 頻道"""
    await bot.wait_until_ready()
    
    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            print("❌ 找不到 Discord 伺服器")
            return
        
        session = Session()
        
        # 查找已確認但還沒有 Discord 頻道的新預約
        new_bookings = session.execute(text("""
            SELECT b.id, b.customerId, b.scheduleId, b."paymentInfo", b.discordDelayMinutes,
                   c.name as customer_name, p.name as partner_name, s.startTime
            FROM "Booking" b
            JOIN "Customer" c ON b.customerId = c.id
            JOIN "Schedule" s ON b.scheduleId = s.id
            JOIN "Partner" p ON s.partnerId = p.id
            WHERE b.status = 'CONFIRMED' 
            AND b.discordEarlyTextChannelId IS NULL
            AND b.createdAt <= NOW() - INTERVAL '3 minutes'
        """)).fetchall()
        
        session.close()
        
        for booking in new_bookings:
            try:
                # 判斷是否為即時預約
                is_instant_booking = False
                if booking.paymentInfo and isinstance(booking.paymentInfo, dict):
                    is_instant_booking = booking.paymentInfo.get('isInstantBooking') == 'true'
                
                # 創建提前溝通文字頻道
                channel = create_booking_text_channel(
                    guild, 
                    booking.id, 
                    booking.customer_name, 
                    booking.partner_name,
                    is_instant_booking
                )
                
                if channel:
                    # 更新資料庫
                    session = Session()
                    session.execute(text("""
                        UPDATE "Booking" 
                        SET "discordEarlyTextChannelId" = :channel_id
                        WHERE id = :booking_id
                    """), {'channel_id': str(channel.id), 'booking_id': booking.id})
                    session.commit()
                    session.close()
                    
                    # 發送歡迎訊息
                    embed = discord.Embed(
                        title="🎮 預約確認",
                        description=f"嗨 {booking.customer_name}！你的預約已確認，夥伴 {booking.partner_name} 將在預約時間與你聯繫。",
                        color=0x00ff00
                    )
                    embed.add_field(name="📅 預約時間", value="請等待夥伴確認具體時間", inline=False)
                    embed.add_field(name="💬 溝通方式", value="此頻道用於預約前的溝通", inline=False)
                    
                    await channel.send(embed=embed)
                    print(f"✅ 創建新預約頻道: {booking.id}")
                
            except Exception as e:
                print(f"❌ 處理新預約 {booking.id} 時發生錯誤: {e}")
                
    except Exception as e:
        print(f"❌ 檢查新預約時發生錯誤: {e}")

# 檢查即時預約的任務
@tasks.loop(seconds=CHECK_INTERVAL)
async def check_instant_bookings_for_voice_channel():
    """檢查即時預約是否需要創建語音頻道"""
    await bot.wait_until_ready()
    
    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            print("❌ 找不到 Discord 伺服器")
            return
        
        session = Session()
        
        # 查找即時預約中需要創建語音頻道的
        now = datetime.now(timezone.utc)
        instant_bookings = session.execute(text("""
            SELECT b.id, b.customerId, b.scheduleId, b.discordEarlyTextChannelId, b.discordVoiceChannelId,
                   c.name as customer_name, p.name as partner_name, s.startTime, b."paymentInfo"
            FROM "Booking" b
            JOIN "Customer" c ON b.customerId = c.id
            JOIN "Schedule" s ON b.scheduleId = s.id
            JOIN "Partner" p ON s.partnerId = p.id
            WHERE b.status = 'CONFIRMED' 
            AND b."paymentInfo"->>'isInstantBooking' = 'true'
            AND b.discordEarlyTextChannelId IS NOT NULL
            AND b.discordVoiceChannelId IS NULL
            AND s.startTime <= :now
        """), {'now': now}).fetchall()
        
        session.close()
        
        for booking in instant_bookings:
            try:
                # 創建語音頻道
                voice_channel = create_booking_voice_channel(
                    guild, 
                    booking.id, 
                    booking.customer_name, 
                    booking.partner_name
                )
                
                if voice_channel:
                    # 創建正式文字頻道
                    text_channel = create_booking_text_channel(
                        guild, 
                        booking.id, 
                        booking.customer_name, 
                        booking.partner_name
                    )
                    
                    if text_channel:
                        # 更新資料庫
                        session = Session()
                        session.execute(text("""
                            UPDATE "Booking" 
                            SET "discordVoiceChannelId" = :voice_id, "discordTextChannelId" = :text_id
                            WHERE id = :booking_id
                        """), {
                            'voice_id': str(voice_channel.id), 
                            'text_id': str(text_channel.id),
                            'booking_id': booking.id
                        })
                        session.commit()
                        session.close()
                        
                        # 刪除提前溝通頻道
                        try:
                            early_channel = guild.get_channel(int(booking.discordEarlyTextChannelId))
                            if early_channel:
                                await early_channel.delete()
                                print(f"✅ 刪除即時預約提前溝通頻道: {booking.id}")
                    except Exception as e:
                            print(f"⚠️ 刪除即時預約提前溝通頻道失敗: {e}")
                        
                        # 在正式文字頻道發送歡迎訊息
                        embed = discord.Embed(
                            title="🎮 即時預約開始",
                            description=f"即時預約已開始！請進入語音頻道開始遊戲。",
                            color=0x0099ff
                        )
                        embed.add_field(name="🎤 語音頻道", value=f"請點擊 {voice_channel.mention} 進入", inline=False)
                        
                        await text_channel.send(embed=embed)
                        print(f"✅ 創建即時預約正式頻道: {booking.id}")
                
                    except Exception as e:
                print(f"❌ 處理即時預約 {booking.id} 時發生錯誤: {e}")
        
                    except Exception as e:
        print(f"❌ 檢查即時預約時發生錯誤: {e}")

# 檢查缺少評價的預約
@tasks.loop(minutes=5)
async def check_missing_ratings():
    """檢查缺少評價的預約並更新狀態"""
    try:
        session = Session()
        
        # 查找已結束但缺少評價的預約
        missing_ratings = session.execute(text("""
            SELECT b.id, s.endTime
            FROM "Booking" b
            JOIN "Schedule" s ON b.scheduleId = s.id
            WHERE b.status = 'CONFIRMED'
            AND s.endTime < NOW() - INTERVAL '1 hour'
            AND NOT EXISTS (
                SELECT 1 FROM "Review" r WHERE r.bookingId = b.id
            )
        """)).fetchall()
        
        # 更新預約狀態為已完成
        for booking in missing_ratings:
            session.execute(text("""
                UPDATE "Booking" 
                SET status = 'COMPLETED'
                WHERE id = :booking_id
            """), {'booking_id': booking.id})
            
            # 計算推薦獎勵
            await calculate_referral_earnings(booking.id)
        
        session.commit()
        session.close()
        
        if missing_ratings:
            print(f"✅ 更新了 {len(missing_ratings)} 個缺少評價的預約狀態")
        
    except Exception as e:
        print(f"❌ 檢查缺少評價時發生錯誤: {e}")

# Bot 事件處理
@bot.event
async def on_ready():
    print(f'✅ {bot.user} 已上線！')
    print(f'📊 伺服器數量: {len(bot.guilds)}')
    
    # 同步 Slash 指令
    try:
        synced = await bot.tree.sync()
        print(f'✅ 已同步 {len(synced)} 個 Slash 指令')
                                        except Exception as e:
        print(f'❌ 同步 Slash 指令失敗: {e}')
    
    # 啟動檢查任務
    if not check_bookings.is_running():
        check_bookings.start()
        print('✅ 啟動預約檢查任務')
    
    if not check_new_bookings.is_running():
        check_new_bookings.start()
        print('✅ 啟動新預約檢查任務')
    
    if not check_instant_bookings_for_voice_channel.is_running():
        check_instant_bookings_for_voice_channel.start()
        print('✅ 啟動即時預約檢查任務')
    
    if not check_missing_ratings.is_running():
        check_missing_ratings.start()
        print('✅ 啟動缺少評價檢查任務')
    
    if not database_health_check.is_running():
        database_health_check.start()
        print('✅ 啟動資料庫健康檢查任務')
    
    if not check_pending_reviews.is_running():
        check_pending_reviews.start()
        print('✅ 啟動待審核檢查任務')

@bot.event
async def on_interaction(interaction):
    """處理所有互動事件"""
    if not interaction.is_component():
        return
    
    custom_id = interaction.custom_id
    
    try:
        if custom_id.startswith('rate_'):
            # 處理評價按鈕
            parts = custom_id.split('_')
            if len(parts) >= 3:
                booking_id = parts[1]
                rating = int(parts[2])
                
                await handle_rating(interaction, booking_id, rating)
        
        elif custom_id.startswith('extend_booking_'):
            # 處理延長預約按鈕
            booking_id = custom_id.replace('extend_booking_', '')
            await handle_extend_booking(interaction, booking_id)
                    
                except Exception as e:
        print(f"❌ 處理互動時發生錯誤: {e}")
        if not interaction.response.is_done():
            await interaction.response.send_message("❌ 處理請求時發生錯誤，請稍後再試。", ephemeral=True)

async def handle_rating(interaction, booking_id, rating):
    """處理評價"""
    try:
        # 更新資料庫中的評價
        session = Session()
        
        # 檢查是否已經評價過
        existing_review = session.execute(text("""
            SELECT id FROM "Review" WHERE bookingId = :booking_id
        """), {'booking_id': booking_id}).fetchone()
        
        if existing_review:
            await interaction.response.send_message("❌ 此預約已經評價過了。", ephemeral=True)
            session.close()
            return
        
        # 獲取預約信息
        booking_info = session.execute(text("""
            SELECT b.customerId, b.scheduleId, s.partnerId
            FROM "Booking" b
            JOIN "Schedule" s ON b.scheduleId = s.id
            WHERE b.id = :booking_id
        """), {'booking_id': booking_id}).fetchone()
        
        if not booking_info:
            await interaction.response.send_message("❌ 找不到預約信息。", ephemeral=True)
            session.close()
            return
        
        # 創建評價記錄
        session.execute(text("""
            INSERT INTO "Review" (id, bookingId, reviewerId, revieweeId, rating, comment, createdAt)
            VALUES (:id, :booking_id, :reviewer_id, :reviewee_id, :rating, :comment, :created_at)
        """), {
            'id': f"review_{booking_id}_{int(time.time())}",
            'booking_id': booking_id,
            'reviewer_id': booking_info.customerId,
            'reviewee_id': booking_info.partnerId,
            'rating': rating,
            'comment': f"自動評價：{rating}星",
            'created_at': datetime.now(timezone.utc)
        })
        
        # 更新預約狀態
        session.execute(text("""
            UPDATE "Booking" 
            SET "ratingCompleted" = true, status = 'COMPLETED'
            WHERE id = :booking_id
        """), {'booking_id': booking_id})
        
        session.commit()
        session.close()
        
        # 發送確認訊息
        stars_display = "⭐" * rating
        embed = discord.Embed(
            title="⭐ 評價完成",
            description=f"感謝你的評價！你給予了 {rating} 顆星評價。",
            color=0x00ff00
        )
        embed.add_field(name="你的評價", value=f"{rating} 顆星 {stars_display}", inline=False)
        embed.add_field(name="評價說明", value="你的評價將幫助其他用戶選擇合適的遊戲夥伴", inline=False)
        
        await interaction.response.send_message(embed=embed, ephemeral=True)
        
        # 通知管理員頻道
        try:
            guild = bot.get_guild(GUILD_ID)
            admin_channel = guild.get_channel(ADMIN_CHANNEL_ID)
            if admin_channel:
                stars_display = "⭐" * rating
                admin_embed = discord.Embed(
                    title="📊 新評價",
                    description=f"預約 {booking_id} 收到新評價",
                    color=0x0099ff
                )
                admin_embed.add_field(name="評價星級", value=f"{rating} 顆星 {stars_display}", inline=True)
                admin_embed.add_field(name="評價時間", value=f"<t:{int(datetime.now(timezone.utc).timestamp())}:F>", inline=True)
                await admin_channel.send(embed=admin_embed)
        except Exception as e:
            print(f"⚠️ 發送管理員通知失敗: {e}")
        
        print(f"✅ 處理評價: {booking_id}, {rating}星")
        
    except Exception as e:
        print(f"❌ 處理評價時發生錯誤: {e}")
        if not interaction.response.is_done():
            await interaction.response.send_message("❌ 評價失敗，請稍後再試。", ephemeral=True)

# --- 發送配對紀錄到管理員頻道（無論是否有評價）---
async def send_pairing_record_to_admin(record_id):
    """發送配對紀錄到管理員頻道（無論是否有評價）"""
    try:
        # 檢查是否已經發送過，避免重複發送
        if record_id in pairing_record_sent:
            print(f"⚠️ 配對紀錄 {record_id} 已經發送過，跳過重複發送")
            return
        
        admin_channel = bot.get_channel(ADMIN_CHANNEL_ID)
        if not admin_channel:
            print(f"❌ 找不到管理員頻道 (ID: {ADMIN_CHANNEL_ID})")
            return
        
        # 從資料庫獲取配對記錄資訊
        session = Session()
        try:
            result = session.execute(text("""
                SELECT "user1Id", "user2Id", duration, "extendedTimes", "bookingId"
                FROM "PairingRecord" 
                WHERE id = :record_id
            """), {"record_id": record_id}).fetchone()
            
            if not result:
                print(f"❌ 找不到配對記錄: {record_id}")
                return
            
            user1_id = result[0]
            user2_id = result[1]
            duration = result[2]
            extended_times = result[3] if result[3] else 0
            booking_id = result[4] if result[4] else None
            
            print(f"🔍 PairingRecord 資訊: record_id={record_id}, user1_id={user1_id}, user2_id={user2_id}, booking_id={booking_id}")
            
            # 驗證用戶ID格式（應該是 Discord ID，通常是 17-19 位數字）
            if not user1_id or not user2_id:
                print(f"⚠️ 警告：PairingRecord {record_id} 中的用戶ID為空")
            elif not user1_id.isdigit() or not user2_id.isdigit():
                print(f"⚠️ 警告：PairingRecord {record_id} 中的用戶ID格式可能錯誤: user1_id={user1_id}, user2_id={user2_id}")
            
            # 如果有 bookingId，從 Booking 獲取正確的 customer 和 partner Discord ID
            # 但如果是 manual_ 前綴，表示這是手動配對，沒有對應的 Booking 記錄，直接使用 PairingRecord 中的用戶ID
            if booking_id:
                if booking_id.startswith('manual_'):
                    print(f"ℹ️ 這是手動配對記錄 (booking_id={booking_id})，直接使用 PairingRecord 中的用戶ID")
                    print(f"✅ 使用 PairingRecord 中的用戶ID: user1_id={user1_id}, user2_id={user2_id}")
                else:
                    print(f"🔍 嘗試從 Booking 獲取用戶資訊: booking_id={booking_id}")
                    
                    booking_result = session.execute(text("""
                        SELECT 
                            c."userId" as customer_user_id,
                            p."userId" as partner_user_id
                        FROM "Booking" b
                        JOIN "Customer" c ON b."customerId" = c.id
                        JOIN "Schedule" s ON b."scheduleId" = s.id
                        JOIN "Partner" p ON s."partnerId" = p.id
                        WHERE b.id = :booking_id
                    """), {"booking_id": booking_id}).fetchone()
                    
                    if booking_result:
                        customer_user_id = booking_result[0]
                        partner_user_id = booking_result[1]
                        print(f"✅ 找到 Booking: customer_user_id={customer_user_id}, partner_user_id={partner_user_id}")
                        
                        # 從 User 表獲取 Discord ID
                        customer_discord_result = session.execute(text("""
                            SELECT discord FROM "User" WHERE id = :user_id
                        """), {"user_id": customer_user_id}).fetchone()
                        
                        partner_discord_result = session.execute(text("""
                            SELECT discord FROM "User" WHERE id = :user_id
                        """), {"user_id": partner_user_id}).fetchone()
                        
                        if customer_discord_result and customer_discord_result[0]:
                            user1_id = customer_discord_result[0]
                            print(f"✅ 更新 user1_id 為: {user1_id}")
                        else:
                            print(f"⚠️ 找不到 customer 的 Discord ID: customer_user_id={customer_user_id}")
                        
                        if partner_discord_result and partner_discord_result[0]:
                            user2_id = partner_discord_result[0]
                            print(f"✅ 更新 user2_id 為: {user2_id}")
                        else:
                            print(f"⚠️ 找不到 partner 的 Discord ID: partner_user_id={partner_user_id}")
                        
                        print(f"🔍 最終 Discord ID: user1_id={user1_id}, user2_id={user2_id}")
                    else:
                        print(f"⚠️ 找不到 Booking 記錄 (booking_id={booking_id})，使用 PairingRecord 中的用戶ID")
                        print(f"⚠️ PairingRecord 中的用戶ID: user1_id={user1_id}, user2_id={user2_id}")
        finally:
            session.close()
        
        # 獲取用戶資訊
        try:
            user1 = await bot.fetch_user(int(user1_id))
            user1_mention = user1.mention
        except:
            user1_mention = f"用戶 {user1_id}"
        
        try:
            user2 = await bot.fetch_user(int(user2_id))
            user2_mention = user2.mention
        except:
            user2_mention = f"用戶 {user2_id}"
        
        # 構建配對紀錄標題
        duration_minutes = duration // 60
        header = f"📋 配對紀錄：{user1_mention} × {user2_mention} | {duration_minutes} 分鐘 | 延長 {extended_times} 次"
        if booking_id:
            header += f"\n預約ID: {booking_id}"
        
        # 檢查是否有評價
        has_ratings = False
        feedback = ""
        
        # 檢查緩存中是否有評價
        if record_id in rating_notification_cache:
            cache_data = rating_notification_cache[record_id]
            ratings = cache_data.get('ratings', [])
            if ratings:
                has_ratings = True
                feedback = "\n⭐ 評價回饋："
                for rating_data in ratings:
                    try:
                        from_user_id = rating_data['user1']
                        to_user_id = rating_data['user2']
                        
                        try:
                            from_user = await bot.fetch_user(int(from_user_id))
                            from_user_mention = from_user.mention
                        except:
                            from_user_mention = f"用戶 {from_user_id}"
                        
                        try:
                            to_user = await bot.fetch_user(int(to_user_id))
                            to_user_mention = to_user.mention
                        except:
                            to_user_mention = f"用戶 {to_user_id}"
                        
                        feedback += f"\n- 「{from_user_mention} → {to_user_mention}」：{rating_data['rating']} ⭐"
                        if rating_data.get('comment'):
                            feedback += f"\n  💬 {rating_data['comment']}"
                    except Exception as e:
                        print(f"⚠️ 處理評價數據時發生錯誤: {e}")
                        continue
        
        # 如果沒有評價，顯示提示
        if not has_ratings:
            feedback = "\n⭐ 沒有收到任何評價。"
        
        # 發送配對紀錄
        await admin_channel.send(f"{header}{feedback}")
        print(f"✅ 配對紀錄已發送到管理員頻道: {record_id}")
        
        # 標記為已發送
        pairing_record_sent.add(record_id)
        
    except Exception as e:
        print(f"❌ 發送配對紀錄到管理員頻道失敗: {e}")
        import traceback
        traceback.print_exc()

# --- 發送合併的評價到管理員頻道（PairingRecord 系統）---
async def send_merged_rating_to_admin(record_id):
    """發送合併的評價結果到管理員頻道"""
    try:
        if record_id not in rating_notification_cache:
            return
        
        cache_data = rating_notification_cache[record_id]
        ratings = cache_data['ratings']
        user1_id = cache_data['user1_id']
        user2_id = cache_data['user2_id']
        
        admin_channel = bot.get_channel(ADMIN_CHANNEL_ID)
        if not admin_channel:
            print(f"❌ 找不到管理員頻道 (ID: {ADMIN_CHANNEL_ID})")
            return
        
        # 獲取用戶資訊
        try:
            user1 = await bot.fetch_user(int(user1_id))
            user1_display = user1.display_name
        except:
            user1_display = f"用戶 {user1_id}"
        
        try:
            user2 = await bot.fetch_user(int(user2_id))
            user2_display = user2.display_name
        except:
            user2_display = f"用戶 {user2_id}"
        
        # 創建合併的評價嵌入訊息
        embed = discord.Embed(
            title="⭐ 新評價回饋",
            color=0x00ff00,
            timestamp=datetime.now(timezone.utc)
        )
        
        # 添加配對記錄ID
        embed.add_field(
            name="📋 配對記錄ID",
            value=f"`{record_id}`",
            inline=False
        )
        
        # 收集所有評價內容
        all_rating_texts = []
        
        for rating_data in ratings:
            try:
                from_user_id = rating_data['user1']
                to_user_id = rating_data['user2']
                role = rating_data.get('role', '未知')
                
                # 獲取評價者和被評價者的顯示名稱
                try:
                    from_user = await bot.fetch_user(int(from_user_id))
                    from_user_display = from_user.display_name
                except:
                    from_user_display = f"用戶 {from_user_id}"
                
                try:
                    to_user = await bot.fetch_user(int(to_user_id))
                    to_user_display = to_user.display_name
                except:
                    to_user_display = f"用戶 {to_user_id}"
                
                # 創建評價區塊（符合圖片格式）
                rating_block = f"👤 **評價者**: {from_user_display}\n"
                rating_block += f"👤 **被評價者**: {to_user_display}\n"
                rating_block += f"⭐ **評分**: {'⭐' * rating_data['rating']}\n"
                rating_block += f"👤 **評價者身份**: {role}\n"
                
                if rating_data.get('comment'):
                    rating_block += f"💬 **留言**: {rating_data['comment']}\n"
                
                all_rating_texts.append(rating_block.strip())
            except Exception as e:
                print(f"⚠️ 處理評價數據時發生錯誤: {e}")
                continue
        
        # 將所有評價合併到一個欄位中（夥伴和顧客會出現在同一欄）
        if all_rating_texts:
            combined_text = "\n\n---\n\n".join(all_rating_texts)
            embed.add_field(
                name="👥 評價內容",
                value=combined_text,
                inline=False
            )
        
        # 格式化時間（轉換為台灣時間並格式化）
        taiwan_time = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=8)))
        hour_12 = taiwan_time.hour % 12
        if hour_12 == 0:
            hour_12 = 12
        time_str = f"{hour_12:02d}:{taiwan_time.minute:02d}"
        am_pm = "上午" if taiwan_time.hour < 12 else "下午"
        footer_text = f"PeiPlay 評價系統・今天{am_pm}{time_str}"
        embed.set_footer(text=footer_text)
        
        await admin_channel.send(embed=embed)
        print(f"✅ 合併評價已發送到管理員頻道: {record_id} ({len(ratings)} 條評價)")
        
        # 清理緩存和計時器
        if record_id in rating_notification_cache:
            cache_data = rating_notification_cache[record_id]
            if cache_data.get('timer') and not cache_data['timer'].done():
                cache_data['timer'].cancel()
            del rating_notification_cache[record_id]
        
    except Exception as e:
        print(f"❌ 發送合併評價到管理員頻道失敗: {e}")
        import traceback
        traceback.print_exc()

# --- 添加評價到緩存並檢查是否需要發送（PairingRecord 系統）---
async def add_rating_to_cache(record_id, rating_data, user1_id, user2_id):
    """添加評價到緩存，如果兩個用戶都已提交則立即發送，否則等待30秒"""
    try:
        # 初始化緩存
        if record_id not in rating_notification_cache:
            rating_notification_cache[record_id] = {
                'ratings': [],
                'user1_id': str(user1_id),
                'user2_id': str(user2_id),
                'timer': None
            }
        
        cache_data = rating_notification_cache[record_id]
        cache_data['ratings'].append(rating_data)
        
        # 檢查是否兩個用戶都已提交評價
        submitted_users = rating_submitted_users.get(record_id, set())
        user1_submitted = str(user1_id) in submitted_users
        user2_submitted = str(user2_id) in submitted_users
        is_single_user = str(user1_id) == str(user2_id)
        
        # 如果兩個用戶都已提交，立即發送
        if (user1_submitted and user2_submitted) or (is_single_user and user1_submitted):
            # 取消現有的計時器（如果有的話）
            if cache_data['timer'] and not cache_data['timer'].done():
                cache_data['timer'].cancel()
            
            # 立即發送
            await send_merged_rating_to_admin(record_id)
        else:
            # 如果只有一個用戶提交，設置30秒計時器
            if cache_data['timer'] is None or cache_data['timer'].done():
                async def delayed_send():
                    await asyncio.sleep(30)  # 等待30秒
                    if record_id in rating_notification_cache:
                        await send_merged_rating_to_admin(record_id)
                
                cache_data['timer'] = asyncio.create_task(delayed_send())
                print(f"⏳ 設置30秒延遲發送評價通知: {record_id}")
        
    except Exception as e:
        print(f"❌ 添加評價到緩存失敗: {e}")
        import traceback
        traceback.print_exc()

# --- 評價選擇 View（包含星等和身份選擇）---
class RatingSelectionView(View):
    """評價選擇界面，包含星等和身份選擇按鈕"""
    def __init__(self, record_id):
        super().__init__(timeout=300)  # 5分鐘超時
        self.record_id = record_id
        self.selected_rating = {}  # {user_id: rating}
        self.selected_role = {}  # {user_id: role}
    
    @discord.ui.button(label="⭐ 1星", style=discord.ButtonStyle.success, emoji="⭐", row=0)
    async def rate_1_star(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_rating[interaction.user.id] = 1
        await interaction.response.send_message("✅ 已選擇1星評分", ephemeral=True)
    
    @discord.ui.button(label="⭐ 2星", style=discord.ButtonStyle.success, emoji="⭐", row=0)
    async def rate_2_star(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_rating[interaction.user.id] = 2
        await interaction.response.send_message("✅ 已選擇2星評分", ephemeral=True)
    
    @discord.ui.button(label="⭐ 3星", style=discord.ButtonStyle.success, emoji="⭐", row=0)
    async def rate_3_star(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_rating[interaction.user.id] = 3
        await interaction.response.send_message("✅ 已選擇3星評分", ephemeral=True)
    
    @discord.ui.button(label="⭐ 4星", style=discord.ButtonStyle.success, emoji="⭐", row=0)
    async def rate_4_star(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_rating[interaction.user.id] = 4
        await interaction.response.send_message("✅ 已選擇4星評分", ephemeral=True)
    
    @discord.ui.button(label="⭐ 5星", style=discord.ButtonStyle.success, emoji="⭐", row=0)
    async def rate_5_star(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_rating[interaction.user.id] = 5
        await interaction.response.send_message("✅ 已選擇5星評分", ephemeral=True)
    
    @discord.ui.button(label="我是顧客", style=discord.ButtonStyle.primary, row=1)
    async def select_customer(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_role[interaction.user.id] = "顧客"
        await interaction.response.send_message("✅ 已選擇身份:顧客", ephemeral=True)
    
    @discord.ui.button(label="我是夥伴", style=discord.ButtonStyle.success, row=1)
    async def select_partner(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.selected_role[interaction.user.id] = "夥伴"
        await interaction.response.send_message("✅ 已選擇身份:夥伴", ephemeral=True)
    
    @discord.ui.button(label="提交評價", style=discord.ButtonStyle.success, row=2)
    async def submit_rating(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            user_id = interaction.user.id
            
            # 檢查是否已選擇評分和身份
            if user_id not in self.selected_rating:
                await interaction.response.send_message("❌ 請先選擇評分（1-5星）", ephemeral=True)
                return
            
            if user_id not in self.selected_role:
                await interaction.response.send_message("❌ 請先選擇身份（顧客或夥伴）", ephemeral=True)
                return
            
            rating = self.selected_rating[user_id]
            role = self.selected_role[user_id]
            
            # 檢查用戶是否已經提交過評價
            if self.record_id in rating_submitted_users:
                if str(user_id) in rating_submitted_users[self.record_id]:
                    await interaction.response.send_message("❗ 您已經提交過評價了。", ephemeral=True)
                    return
            
            # 打開留言表單（選填）
            modal = RatingCommentModal(self.record_id, rating, role)
            await interaction.response.send_modal(modal)
            
        except Exception as e:
            print(f"❌ 處理評價提交時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            try:
                if not interaction.response.is_done():
                    await interaction.response.send_message("❌ 處理請求時發生錯誤，請稍後再試。", ephemeral=True)
            except:
                pass

# --- 留言 Modal（選填）---
class RatingCommentModal(Modal, title="匿名評分與留言"):
    def __init__(self, record_id, rating, role):
        super().__init__()
        self.record_id = record_id
        self.rating = rating
        self.role = role
        
        # 顯示已選擇的評分（只讀）
        self.rating_display = TextInput(
            label="評分",
            default=f"{'⭐' * rating} ({rating} 星)",
            style=discord.TextStyle.short,
            required=False,
            max_length=20
        )
        self.rating_display.disabled = True  # 設為只讀
        self.add_item(self.rating_display)
        
        # 顯示已選擇的身份（只讀）
        self.role_display = TextInput(
            label="身份",
            default=role,
            style=discord.TextStyle.short,
            required=False,
            max_length=10
        )
        self.role_display.disabled = True  # 設為只讀
        self.add_item(self.role_display)
        
        # 留言輸入框
        self.comment = TextInput(
            label="留下你的留言（選填）",
            required=False,
            style=discord.TextStyle.paragraph,
            placeholder="可以留下您的意見或建議...",
            max_length=4000
        )
        self.add_item(self.comment)

    async def on_submit(self, interaction: discord.Interaction):
        try:
            comment_text = self.comment.value.strip() if self.comment.value else ""
            
            print(f"🔍 收到評價提交: record_id={self.record_id}, rating={self.rating}, role={self.role}, comment={comment_text}")
            
            # 使用新的 session 來避免連接問題
            session = Session()
            try:
                # 查詢 PairingRecord（需要根據實際的資料庫結構調整）
                result = session.execute(text("""
                    SELECT "user1Id", "user2Id" 
                    FROM "PairingRecord" 
                    WHERE id = :record_id
                """), {"record_id": self.record_id}).fetchone()
                
                if not result:
                    print(f"❌ 找不到配對記錄: {self.record_id}")
                    await interaction.response.send_message("❌ 找不到配對記錄", ephemeral=True)
                    return
                
                user1_id = result[0]
                user2_id = result[1]
                
                # 更新評價到資料庫
                session.execute(text("""
                    UPDATE "PairingRecord" 
                    SET rating = :rating, comment = :comment
                    WHERE id = :record_id
                """), {
                    "record_id": self.record_id,
                    "rating": self.rating,
                    "comment": f"[{self.role}] {comment_text}" if comment_text else f"[{self.role}]"
                })
                session.commit()
            finally:
                session.close()
            
            await interaction.response.send_message("✅ 感謝你的匿名評價！", ephemeral=True)

            # 標記用戶已提交評價（統一使用字符串格式）
            if self.record_id not in rating_submitted_users:
                rating_submitted_users[self.record_id] = set()
            rating_submitted_users[self.record_id].add(str(interaction.user.id))

            if self.record_id not in pending_ratings:
                pending_ratings[self.record_id] = []
            
            rating_data = {
                'rating': self.rating,
                'role': self.role,
                'comment': comment_text,
                'user1': str(interaction.user.id),
                'user2': str(user2_id if str(interaction.user.id) == user1_id else user1_id)
            }
            pending_ratings[self.record_id].append(rating_data)
            print(f"✅ 評價已添加到待處理列表: {rating_data}")

            # 添加到緩存並檢查是否需要發送合併的通知
            await add_rating_to_cache(self.record_id, rating_data, user1_id, user2_id)

            evaluated_records.add(self.record_id)
            print(f"✅ 評價流程完成")
            
            # 檢查是否所有用戶都已提交評價，如果是則刪除文字頻道
            if self.record_id in rating_text_channels:
                text_channel = rating_text_channels[self.record_id]
                
                # 檢查是否所有相關用戶都已提交
                session = Session()
                try:
                    result = session.execute(text("""
                        SELECT "user1Id", "user2Id" 
                        FROM "PairingRecord" 
                        WHERE id = :record_id
                    """), {"record_id": self.record_id}).fetchone()
                    
                    if result:
                        user1_id = result[0]
                        user2_id = result[1]
                        
                        submitted_users = rating_submitted_users.get(self.record_id, set())
                        
                        # 檢查兩個用戶是否都已提交評價（統一使用字符串格式比較）
                        user1_submitted = str(user1_id) in submitted_users
                        user2_submitted = str(user2_id) in submitted_users
                        
                        # 檢查是否只有一個用戶（自己配對自己）
                        is_single_user = str(user1_id) == str(user2_id)
                        
                        # 如果兩個用戶都已提交，或者只有一個用戶且已提交，則刪除頻道
                        if (user1_submitted and user2_submitted) or (is_single_user and user1_submitted):
                            try:
                                if text_channel and not text_channel.deleted:
                                    await text_channel.delete()
                                    print(f"✅ 所有用戶已提交評價，已刪除文字頻道: {text_channel.name}")
                                    # 清理追蹤
                                    rating_text_channels.pop(self.record_id, None)
                                    rating_channel_created_time.pop(self.record_id, None)
                                    
                                    # 發送配對紀錄到管理員頻道（無論是否有評價）
                                    await send_pairing_record_to_admin(self.record_id)
                            except Exception as e:
                                print(f"❌ 刪除文字頻道失敗: {e}")
                finally:
                    session.close()
        except Exception as e:
            print(f"❌ 評分提交錯誤: {e}")
            import traceback
            traceback.print_exc()
            try:
                await interaction.response.send_message("❌ 提交失敗，請稍後再試", ephemeral=True)
            except:
                # 如果已經回應過，就忽略錯誤
                pass

# --- 評價按鈕 View ---
class RatingSubmitButton(View):
    """評價提交按鈕，點擊後會打開評價選擇界面"""
    def __init__(self, record_id):
        super().__init__(timeout=None)  # 設置為 None，讓按鈕永久有效
        self.record_id = record_id

    @discord.ui.button(label="⭐ 匿名評分", style=discord.ButtonStyle.success, emoji="⭐")
    async def submit(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            print(f"🔍 用戶 {interaction.user.id} 點擊了評價按鈕，record_id={self.record_id}")
            
            # 檢查用戶是否已經提交過評價（使用全局字典，統一使用字符串格式）
            if self.record_id in rating_submitted_users:
                if str(interaction.user.id) in rating_submitted_users[self.record_id]:
                    await interaction.response.send_message("❗ 您已經提交過評價了。", ephemeral=True)
                    return
            
            # 打開評價選擇界面（包含星等和身份選擇按鈕）
            rating_view = RatingSelectionView(self.record_id)
            await interaction.response.send_message(
                "📝 請選擇您的評分和身份，然後點擊「提交評價」按鈕：",
                view=rating_view,
                ephemeral=True
            )
            
        except Exception as e:
            print(f"❌ 處理評價按鈕點擊時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            try:
                if not interaction.response.is_done():
                    await interaction.response.send_message("❌ 處理請求時發生錯誤，請稍後再試。", ephemeral=True)
            except:
                pass

async def handle_extend_booking(interaction, booking_id):
    """處理延長預約"""
    try:
        # 更新資料庫中的預約時間
        session = Session()
        
        # 獲取當前結束時間
        current_end_time = session.execute(text("""
            SELECT s.endTime FROM "Booking" b
            JOIN "Schedule" s ON b.scheduleId = s.id
            WHERE b.id = :booking_id
        """), {'booking_id': booking_id}).fetchone()
        
        if not current_end_time:
            await interaction.response.send_message("❌ 找不到預約信息。", ephemeral=True)
            session.close()
            return
        
                # 延長5分鐘
        new_end_time = current_end_time.endTime + timedelta(minutes=5)
        
        # 更新結束時間
        session.execute(text("""
                    UPDATE "Schedule" 
            SET "endTime" = :new_end_time
                    WHERE id = (
                SELECT scheduleId FROM "Booking" WHERE id = :booking_id
            )
        """), {'new_end_time': new_end_time, 'booking_id': booking_id})
        
        session.commit()
        session.close()
            
            # 發送確認訊息
        embed = discord.Embed(
            title="⏰ 預約已延長",
            description=f"預約已延長 5 分鐘，新的結束時間是 <t:{int(new_end_time.timestamp())}:F>",
            color=0x00ff00
        )
        
        await interaction.response.send_message(embed=embed)
        
        print(f"✅ 延長預約: {booking_id}, 新結束時間: {new_end_time}")
            
        except Exception as e:
        print(f"❌ 延長預約時發生錯誤: {e}")
        if not interaction.response.is_done():
            await interaction.response.send_message("❌ 延長預約失敗，請稍後再試。", ephemeral=True)

# Slash 指令
@bot.tree.command(name="ping", description="檢查 Bot 延遲")
async def ping(interaction: discord.Interaction):
    """檢查 Bot 延遲"""
    latency = round(bot.latency * 1000)
    embed = discord.Embed(
        title="🏓 Pong!",
        description=f"延遲: {latency}ms",
        color=0x00ff00
    )
    await interaction.response.send_message(embed=embed, ephemeral=True)

@bot.tree.command(name="status", description="檢查 Bot 狀態")
async def status(interaction: discord.Interaction):
    """檢查 Bot 狀態"""
    try:
        # 檢查資料庫連線
        session = Session()
        session.execute(text("SELECT 1"))
        db_status = "✅ 正常"
        session.close()
    except:
        db_status = "❌ 異常"
    
    # 檢查任務狀態
    tasks_status = []
    if check_bookings.is_running():
        tasks_status.append("✅ 預約檢查")
    else:
        tasks_status.append("❌ 預約檢查")
    
    if check_new_bookings.is_running():
        tasks_status.append("✅ 新預約檢查")
                    else:
        tasks_status.append("❌ 新預約檢查")
    
    if database_health_check.is_running():
        tasks_status.append("✅ 資料庫健康檢查")
    else:
        tasks_status.append("❌ 資料庫健康檢查")
    
    embed = discord.Embed(
        title="📊 Bot 狀態",
        color=0x0099ff
    )
    embed.add_field(name="資料庫連線", value=db_status, inline=False)
    embed.add_field(name="任務狀態", value="\n".join(tasks_status), inline=False)
    embed.add_field(name="伺服器數量", value=str(len(bot.guilds)), inline=True)
    embed.add_field(name="延遲", value=f"{round(bot.latency * 1000)}ms", inline=True)
    
    await interaction.response.send_message(embed=embed, ephemeral=True)

@bot.tree.command(name="cleanup", description="清理孤立的 Discord 頻道")
async def cleanup(interaction: discord.Interaction):
    """清理孤立的 Discord 頻道"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
            return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            await interaction.followup.send("❌ 找不到 Discord 伺服器")
            return
        
        # 獲取所有預約頻道
        session = Session()
        all_channels = session.execute(text("""
            SELECT "discordEarlyTextChannelId", "discordTextChannelId", "discordVoiceChannelId"
            FROM "Booking"
            WHERE "discordEarlyTextChannelId" IS NOT NULL 
            OR "discordTextChannelId" IS NOT NULL 
            OR "discordVoiceChannelId" IS NOT NULL
        """)).fetchall()
        session.close()
        
        # 收集所有有效的頻道 ID
        valid_channel_ids = set()
        for channel in all_channels:
            if channel.discordEarlyTextChannelId:
                valid_channel_ids.add(int(channel.discordEarlyTextChannelId))
            if channel.discordTextChannelId:
                valid_channel_ids.add(int(channel.discordTextChannelId))
            if channel.discordVoiceChannelId:
                valid_channel_ids.add(int(channel.discordVoiceChannelId))
        
        # 檢查所有頻道
            deleted_count = 0
        for channel in guild.channels:
            # 檢查是否是預約頻道（包含特殊字符）
            if any(char in channel.name for char in ['📝', '🎤', '🔥']):
                if channel.id not in valid_channel_ids:
                try:
                    await channel.delete()
                    deleted_count += 1
                        print(f"✅ 刪除孤立頻道: {channel.name}")
                except Exception as e:
                    print(f"❌ 刪除頻道失敗 {channel.name}: {e}")
            
        embed = discord.Embed(
            title="🧹 清理完成",
            description=f"已刪除 {deleted_count} 個孤立的頻道",
            color=0x00ff00
        )
        await interaction.followup.send(embed=embed)
            
    except Exception as e:
        print(f"❌ 清理頻道時發生錯誤: {e}")
        await interaction.followup.send("❌ 清理失敗，請稍後再試。")

@bot.tree.command(name="force_cleanup", description="強制清理所有預約頻道")
async def force_cleanup(interaction: discord.Interaction):
    """強制清理所有預約頻道"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            await interaction.followup.send("❌ 找不到 Discord 伺服器")
                return
                
        deleted_count = 0
        for channel in guild.channels:
            # 檢查是否是預約頻道
            if any(char in channel.name for char in ['📝', '🎤', '🔥']):
                try:
                    await channel.delete()
                    deleted_count += 1
                    print(f"✅ 強制刪除頻道: {channel.name}")
        except Exception as e:
                    print(f"❌ 強制刪除頻道失敗 {channel.name}: {e}")
        
        embed = discord.Embed(
            title="🧹 強制清理完成",
            description=f"已刪除 {deleted_count} 個預約頻道",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed)
        
        except Exception as e:
        print(f"❌ 強制清理頻道時發生錯誤: {e}")
        await interaction.followup.send("❌ 強制清理失敗，請稍後再試。")

@bot.tree.command(name="emergency_cleanup", description="緊急清理所有頻道")
async def emergency_cleanup(interaction: discord.Interaction):
    """緊急清理所有頻道"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            await interaction.followup.send("❌ 找不到 Discord 伺服器")
            return
        
        deleted_count = 0
        for channel in guild.channels:
            # 刪除所有非系統頻道
            if channel.type in [discord.ChannelType.text, discord.ChannelType.voice]:
                try:
                    await channel.delete()
                    deleted_count += 1
                    print(f"✅ 緊急刪除頻道: {channel.name}")
        except Exception as e:
                    print(f"❌ 緊急刪除頻道失敗 {channel.name}: {e}")
        
        embed = discord.Embed(
            title="🚨 緊急清理完成",
            description=f"已刪除 {deleted_count} 個頻道",
            color=0xff0000
        )
        await interaction.followup.send(embed=embed)
        
        except Exception as e:
        print(f"❌ 緊急清理頻道時發生錯誤: {e}")
        await interaction.followup.send("❌ 緊急清理失敗，請稍後再試。")

@bot.tree.command(name="stats", description="顯示預約統計")
async def stats(interaction: discord.Interaction):
    """顯示預約統計"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        session = Session()
        
        # 獲取統計數據
        total_bookings = session.execute(text("SELECT COUNT(*) FROM \"Booking\"")).fetchone()[0]
        confirmed_bookings = session.execute(text("SELECT COUNT(*) FROM \"Booking\" WHERE status = 'CONFIRMED'")).fetchone()[0]
        completed_bookings = session.execute(text("SELECT COUNT(*) FROM \"Booking\" WHERE status = 'COMPLETED'")).fetchone()[0]
        
        # 獲取今天的預約
        today = datetime.now(timezone.utc).date()
        today_bookings = session.execute(text("""
            SELECT COUNT(*) FROM "Booking" b
            JOIN "Schedule" s ON b.scheduleId = s.id
            WHERE DATE(s.startTime) = :today
        """), {'today': today}).fetchone()[0]
        
        # 獲取活躍頻道
        active_channels = session.execute(text("""
            SELECT COUNT(*) FROM "Booking"
            WHERE "discordEarlyTextChannelId" IS NOT NULL
            OR "discordTextChannelId" IS NOT NULL
            OR "discordVoiceChannelId" IS NOT NULL
        """)).fetchone()[0]
        
        session.close()
        
            embed = discord.Embed(
            title="📊 預約統計",
            color=0x0099ff
        )
        embed.add_field(name="總預約數", value=str(total_bookings), inline=True)
        embed.add_field(name="已確認預約", value=str(confirmed_bookings), inline=True)
        embed.add_field(name="已完成預約", value=str(completed_bookings), inline=True)
        embed.add_field(name="今日預約", value=str(today_bookings), inline=True)
        embed.add_field(name="活躍頻道", value=str(active_channels), inline=True)
        embed.add_field(name="檢查間隔", value=f"{CHECK_INTERVAL}秒", inline=True)
        
        await interaction.followup.send(embed=embed)
            
        except Exception as e:
        print(f"❌ 獲取統計數據時發生錯誤: {e}")
        await interaction.followup.send("❌ 獲取統計數據失敗，請稍後再試。")

@bot.tree.command(name="test_notification", description="測試管理員通知")
async def test_notification(interaction: discord.Interaction):
    """測試管理員通知"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
                    return
    
    try:
        guild = bot.get_guild(GUILD_ID)
        admin_channel = guild.get_channel(ADMIN_CHANNEL_ID)
        
        if admin_channel:
            embed = discord.Embed(
                title="🧪 測試通知",
                description="這是一個測試通知，用於確認管理員頻道正常工作。",
                color=0x0099ff
            )
            await admin_channel.send(embed=embed)
            await interaction.response.send_message("✅ 測試通知已發送", ephemeral=True)
                else:
            await interaction.response.send_message("❌ 找不到管理員頻道", ephemeral=True)
            
    except Exception as e:
        print(f"❌ 發送測試通知時發生錯誤: {e}")
        await interaction.response.send_message("❌ 發送測試通知失敗", ephemeral=True)

@bot.tree.command(name="debug_booking", description="調試特定預約")
async def debug_booking(interaction: discord.Interaction, booking_id: str):
    """調試特定預約"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)
    
    try:
        session = Session()
        
        # 獲取預約詳細信息
        booking_info = session.execute(text("""
            SELECT b.*, c.name as customer_name, p.name as partner_name, 
                   s.startTime, s.endTime, s.status as schedule_status
            FROM "Booking" b
            JOIN "Customer" c ON b.customerId = c.id
            JOIN "Schedule" s ON b.scheduleId = s.id
            JOIN "Partner" p ON s.partnerId = p.id
            WHERE b.id = :booking_id
        """), {'booking_id': booking_id}).fetchone()
        
        session.close()
        
        if not booking_info:
            await interaction.followup.send("❌ 找不到指定的預約")
            return
        
        # 創建調試信息嵌入
        embed = discord.Embed(
            title=f"🔍 預約調試: {booking_id}",
            color=0x0099ff
        )
        
        embed.add_field(name="客戶", value=booking_info.customer_name, inline=True)
        embed.add_field(name="夥伴", value=booking_info.partner_name, inline=True)
        embed.add_field(name="狀態", value=booking_info.status, inline=True)
        embed.add_field(name="開始時間", value=f"<t:{int(booking_info.startTime.timestamp())}:F>", inline=True)
        embed.add_field(name="結束時間", value=f"<t:{int(booking_info.endTime.timestamp())}:F>", inline=True)
        embed.add_field(name="時程狀態", value=booking_info.schedule_status, inline=True)
        
        # Discord 頻道信息
        discord_info = []
        if booking_info.discordEarlyTextChannelId:
            discord_info.append(f"提前文字: {booking_info.discordEarlyTextChannelId}")
        if booking_info.discordTextChannelId:
            discord_info.append(f"正式文字: {booking_info.discordTextChannelId}")
        if booking_info.discordVoiceChannelId:
            discord_info.append(f"語音: {booking_info.discordVoiceChannelId}")
        
        if discord_info:
            embed.add_field(name="Discord 頻道", value="\n".join(discord_info), inline=False)
        
        # 其他信息
        is_instant_booking = False
        if booking_info.paymentInfo and isinstance(booking_info.paymentInfo, dict):
            is_instant_booking = booking_info.paymentInfo.get('isInstantBooking') == 'true'
        embed.add_field(name="即時預約", value="是" if is_instant_booking else "否", inline=True)
        embed.add_field(name="延長按鈕", value="已顯示" if booking_info.extensionButtonShown else "未顯示", inline=True)
        embed.add_field(name="評價完成", value="是" if booking_info.ratingCompleted else "否", inline=True)
        
        await interaction.followup.send(embed=embed)
        
    except Exception as e:
        print(f"❌ 調試預約時發生錯誤: {e}")
        await interaction.followup.send("❌ 調試預約失敗，請稍後再試。")

@bot.tree.command(name="export_pairing_records", description="導出配對記錄到Excel")
async def export_pairing_records(interaction: discord.Interaction):
    """導出配對記錄到Excel文件"""
    if not interaction.user.id == ADMIN_USER_ID:
        await interaction.response.send_message("❌ 只有管理員可以使用此指令。", ephemeral=True)
        return
    
    await interaction.response.defer(ephemeral=True)
    
    try:
        session = Session()
        
        # 獲取所有配對記錄
        pairing_records = session.execute(text("""
            SELECT pr.id, pr."user1Id", pr."user2Id", pr.timestamp, pr.duration, pr."bookingId"
            FROM "PairingRecord" pr
            ORDER BY pr.timestamp DESC
        """)).fetchall()
        
        if not pairing_records:
            await interaction.followup.send("❌ 沒有找到配對記錄")
            session.close()
            return
        
        # 獲取所有相關的Booking信息
        booking_ids = [r.bookingId for r in pairing_records if r.bookingId and not r.bookingId.startswith('manual_')]
        
        bookings_map = {}
        if booking_ids:
            bookings = session.execute(text("""
                SELECT b.id, 
                       c."userId" as customer_user_id,
                       p."userId" as partner_user_id
                FROM "Booking" b
                JOIN "Customer" c ON b."customerId" = c.id
                JOIN "Schedule" s ON b."scheduleId" = s.id
                JOIN "Partner" p ON s."partnerId" = p.id
                WHERE b.id = ANY(:booking_ids)
            """), {"booking_ids": booking_ids}).fetchall()
            
            for booking in bookings:
                bookings_map[booking.id] = {
                    'customer_user_id': booking.customer_user_id,
                    'partner_user_id': booking.partner_user_id
                }
        
        # 獲取所有用戶的Discord信息
        user_ids = set()
        for booking in bookings_map.values():
            user_ids.add(booking['customer_user_id'])
            user_ids.add(booking['partner_user_id'])
        
        discord_map = {}
        if user_ids:
            users = session.execute(text("""
                SELECT id, discord FROM "User" WHERE id = ANY(:user_ids)
            """), {"user_ids": list(user_ids)}).fetchall()
            
            for user in users:
                discord_map[user.id] = user.discord
        
        session.close()
        
        # 處理配對記錄數據
        records_data = []
        for record in pairing_records:
            booking = bookings_map.get(record.bookingId) if record.bookingId else None
            
            partner_discord = ''
            customer_discord = ''
            
            if booking:
                # 從Booking獲取正確的伙伴和顧客信息
                partner_user_id = booking['partner_user_id']
                customer_user_id = booking['customer_user_id']
                partner_discord = discord_map.get(partner_user_id, '')
                customer_discord = discord_map.get(customer_user_id, '')
            else:
                # 如果沒有bookingId，跳過（無法確定誰是伙伴誰是顧客）
                continue
            
            if not partner_discord or not customer_discord:
                continue
            
            # 轉換時間戳為台灣時間
            timestamp = record.timestamp
            if timestamp.tzinfo is None:
                timestamp = timestamp.replace(tzinfo=timezone.utc)
            tw_time = timestamp.astimezone(timezone(timedelta(hours=8)))
            
            date_str = tw_time.strftime('%Y-%m-%d')
            time_str = tw_time.strftime('%H:%M:%S')
            
            records_data.append({
                'date': date_str,
                'time': time_str,
                'duration': record.duration,
                'partner_discord': partner_discord,
                'customer_discord': customer_discord,
                'timestamp': tw_time
            })
        
        if not records_data:
            await interaction.followup.send("❌ 沒有有效的配對記錄可以導出")
            return
        
        # 按伙伴Discord名字分组
        records_by_partner = {}
        for record in records_data:
            partner_key = record['partner_discord']
            if partner_key not in records_by_partner:
                records_by_partner[partner_key] = []
            records_by_partner[partner_key].append(record)
        
        # 按伙伴Discord名字排序
        sorted_partners = sorted(records_by_partner.keys())
        
        # 創建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "訂單記錄"
        
        # 設置列標題
        headers = ['日期', '時間', '時長(分鐘)', '夥伴 Discord 名字', '顧客 Discord 名字']
        ws.append(headers)
        
        # 設置標題行樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 按伙伴分組添加數據
        for partner_key in sorted_partners:
            partner_records = records_by_partner[partner_key]
            
            # 按時間排序（最新的在前）
            partner_records.sort(key=lambda x: x['timestamp'], reverse=True)
            
            # 添加該伙伴的所有記錄
            for record in partner_records:
                ws.append([
                    record['date'],
                    record['time'],
                    record['duration'],
                    record['partner_discord'],
                    record['customer_discord']
                ])
            
            # 在不同伙伴之間添加空行
            if sorted_partners.index(partner_key) < len(sorted_partners) - 1:
                ws.append([])
        
        # 調整列寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        
        # 保存到內存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 生成文件名
        tw_now = datetime.now(timezone(timedelta(hours=8)))
        filename = f"訂單記錄_{tw_now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 創建Discord文件對象
        file = discord.File(excel_buffer, filename=filename)
        
        # 發送到當前頻道
        await interaction.followup.send("📊 **訂單記錄已更新**", file=file)
        
        print(f"✅ 配對記錄Excel已生成並發送: {filename}")
        
    except Exception as e:
        print(f"❌ 導出配對記錄時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        await interaction.followup.send(f"❌ 導出失敗: {str(e)}")

# Flask API 設定
app = Flask(__name__)

@app.route('/create_instant_text_channel', methods=['POST'])
def create_instant_text_channel():
    """為即時預約創建文字頻道"""
    try:
    data = request.get_json()
        booking_id = data.get('booking_id')
        customer_name = data.get('customer_name')
        partner_name = data.get('partner_name')
        
        if not all([booking_id, customer_name, partner_name]):
            return jsonify({'error': '缺少必要參數'}), 400
        
        # 獲取 Discord 伺服器
            guild = bot.get_guild(GUILD_ID)
            if not guild:
            return jsonify({'error': '找不到 Discord 伺服器'}), 500
        
        # 創建文字頻道
        channel = create_booking_text_channel(guild, booking_id, customer_name, partner_name, True)
        
        if channel:
            # 更新資料庫
            session = Session()
            session.execute(text("""
                UPDATE "Booking" 
                SET "discordEarlyTextChannelId" = :channel_id
                WHERE id = :booking_id
            """), {'channel_id': str(channel.id), 'booking_id': booking_id})
            session.commit()
            session.close()
            
            return jsonify({
                'success': True,
                'channel_id': str(channel.id),
                'channel_name': channel.name
            })
                else:
            return jsonify({'error': '創建頻道失敗'}), 500
            
    except Exception as e:
        print(f"❌ 創建即時文字頻道時發生錯誤: {e}")
        return jsonify({'error': '創建頻道時發生錯誤'}), 500

@app.route('/invite_user', methods=['POST'])
def invite_user():
    """邀請用戶加入 Discord 伺服器"""
    try:
        data = request.get_json()
        discord_name = data.get('discord_name')
        user_name = data.get('user_name')
        user_email = data.get('user_email')
        
        if not discord_name:
            return jsonify({'error': '缺少 Discord 名稱'}), 400
        
        # 獲取 Discord 伺服器
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            return jsonify({'error': '找不到 Discord 伺服器'}), 500
        
        # 嘗試找到用戶
        member = None
        for m in guild.members:
            if m.display_name == discord_name or m.name == discord_name:
                member = m
                break
        
        if member:
            # 發送歡迎訊息
            embed = discord.Embed(
                title="🎉 歡迎加入 PeiPlay！",
                description=f"嗨 {user_name}！歡迎加入我們的 Discord 伺服器！",
                color=0x00ff00
            )
            embed.add_field(name="📧 註冊信箱", value=user_email, inline=False)
            embed.add_field(name="💡 使用提示", value="你可以在這裡找到遊戲夥伴，預約陪玩服務！", inline=False)
            
            try:
                await member.send(embed=embed)
                return jsonify({'success': True, 'message': '歡迎訊息已發送'})
            except:
                return jsonify({'success': True, 'message': '用戶已找到，但無法發送私訊'})
            else:
            # 通知管理員
            admin_channel = guild.get_channel(ADMIN_CHANNEL_ID)
            if admin_channel:
                embed = discord.Embed(
                    title="👤 新用戶註冊",
                    description=f"新用戶 {user_name} 註冊了 PeiPlay，但找不到 Discord 用戶 {discord_name}",
                    color=0xff9900
                )
                embed.add_field(name="📧 信箱", value=user_email, inline=False)
                embed.add_field(name="🎮 Discord 名稱", value=discord_name, inline=False)
                await admin_channel.send(embed=embed)
            
            return jsonify({'success': False, 'message': '找不到 Discord 用戶，已通知管理員'})
            
    except Exception as e:
        print(f"❌ 邀請用戶時發生錯誤: {e}")
        return jsonify({'error': '邀請用戶時發生錯誤'}), 500

@app.route('/delete-channel', methods=['POST'])
def delete_channel():
    """刪除 Discord 頻道"""
    try:
        data = request.get_json()
        channel_id = data.get('channelId')
        
        if not channel_id:
            return jsonify({'error': '缺少 channelId 參數'}), 400
        
        # 獲取 Discord 伺服器
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            return jsonify({'error': '找不到 Discord 伺服器'}), 500
        
        # 刪除頻道
        async def delete_channel_async():
            try:
                channel = guild.get_channel(int(channel_id))
                if channel:
                    await channel.delete()
                    print(f"✅ 已刪除頻道: {channel.name} ({channel_id})")
                    return True
                else:
                    print(f"⚠️ 找不到頻道: {channel_id}")
                    return False
            except Exception as e:
                print(f"❌ 刪除頻道失敗: {e}")
                return False
        
        # 在事件循環中執行異步操作
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(delete_channel_async())
        loop.close()
        
        if result:
            return jsonify({'success': True, 'message': '頻道已刪除'})
        else:
            return jsonify({'success': False, 'message': '找不到頻道或刪除失敗'}), 404
            
    except Exception as e:
        print(f"❌ 刪除頻道時發生錯誤: {e}")
        return jsonify({'error': '刪除頻道時發生錯誤'}), 500

@app.route('/create-group-text-channel', methods=['POST'])
def create_group_text_channel():
    """創建群組文字頻道（匿名文字區）"""
    try:
        data = request.get_json()
        group_id = data.get('groupId')
        group_title = data.get('groupTitle', '')
        participants = data.get('participants', [])
        start_time = data.get('startTime')
        
        if not group_id:
            return jsonify({'error': '缺少 groupId 參數'}), 400
        
        # 獲取 Discord 伺服器
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            return jsonify({'error': '找不到 Discord 伺服器'}), 500
        
        # 創建頻道
        async def create_channel_async():
            try:
                # 設置權限
                overwrites = {
                    guild.default_role: discord.PermissionOverwrite(view_channel=False),
                }
                
                # 為參與者設置權限（排除管理員）
                non_admin_participants = []
                for participant_discord_id in participants:
                    if participant_discord_id:
                        # 排除管理員帳號
                        if str(participant_discord_id) == str(ADMIN_USER_ID):
                            continue
                        member = guild.get_member(int(participant_discord_id))
                        if member:
                            overwrites[member] = discord.PermissionOverwrite(
                                view_channel=True,
                                send_messages=True,
                                read_message_history=True
                            )
                            non_admin_participants.append(participant_discord_id)
                
                # 找到分類
                category = discord.utils.get(guild.categories, name="語音頻道")
                if not category:
                    category = discord.utils.get(guild.categories, name="Voice Channels")
                
                # 創建文字頻道
                text_channel = await guild.create_text_channel(
                    name="🔒匿名文字區",
                    overwrites=overwrites,
                    category=category
                )
                
                # 發送歡迎訊息
                embed = discord.Embed(
                    title="歡迎來到 # 匿名文字區!",
                    description=f"這就是 # 匿名文字區 私人頻道的起點。",
                    color=0x00ff00
                )
                if start_time:
                    try:
                        start_dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                        embed.add_field(
                            name="預約時間",
                            value=f"<t:{int(start_dt.timestamp())}:F>",
                            inline=False
                        )
                    except:
                        pass
                
                await text_channel.send(embed=embed)
                
                # 發送邀請成員消息（排除管理員）
                if non_admin_participants:
                    mentions = " ".join([f"<@{pid}>" for pid in non_admin_participants])
                    await text_channel.send(f"👥 邀請成員：{mentions}")
                
                print(f"✅ 已創建群組文字頻道: {text_channel.name} ({text_channel.id})")
                return text_channel.id
            except Exception as e:
                print(f"❌ 創建群組文字頻道失敗: {e}")
                return None
        
        # 在事件循環中執行異步操作
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        channel_id = loop.run_until_complete(create_channel_async())
        loop.close()
        
        if channel_id:
            return jsonify({
                'success': True,
                'channelId': str(channel_id)
            })
        else:
            return jsonify({'error': '創建頻道失敗'}), 500
            
    except Exception as e:
        print(f"❌ 創建群組文字頻道時發生錯誤: {e}")
        return jsonify({'error': '創建頻道時發生錯誤'}), 500

@app.route('/create-group-voice-channel', methods=['POST'])
def create_group_voice_channel():
    """創建群組語音頻道"""
    try:
        data = request.get_json()
        group_id = data.get('groupId')
        group_title = data.get('groupTitle', '')
        participants = data.get('participants', [])
        start_time = data.get('startTime')
        
        if not group_id:
            return jsonify({'error': '缺少 groupId 參數'}), 400
        
        # 獲取 Discord 伺服器
        guild = bot.get_guild(GUILD_ID)
        if not guild:
            return jsonify({'error': '找不到 Discord 伺服器'}), 500
        
        # 創建頻道
        async def create_channel_async():
            try:
                # 設置權限
                overwrites = {
                    guild.default_role: discord.PermissionOverwrite(view_channel=False),
                }
                
                # 為參與者設置權限（排除管理員）
                non_admin_participants = []
                for participant_discord_id in participants:
                    if participant_discord_id:
                        # 排除管理員帳號
                        if str(participant_discord_id) == str(ADMIN_USER_ID):
                            continue
                        member = guild.get_member(int(participant_discord_id))
                        if member:
                            overwrites[member] = discord.PermissionOverwrite(
                                view_channel=True,
                                connect=True
                            )
                            non_admin_participants.append(participant_discord_id)
                
                # 找到分類
                category = discord.utils.get(guild.categories, name="語音頻道")
                if not category:
                    category = discord.utils.get(guild.categories, name="Voice Channels")
                
                # 生成頻道名稱
                channel_name = group_title if group_title else f"群組語音頻道-{group_id[:8]}"
                
                # 創建語音頻道
                voice_channel = await guild.create_voice_channel(
                    name=channel_name,
                    overwrites=overwrites,
                    category=category
                )
                
                # 找到對應的文字頻道並發送邀請成員消息（排除管理員）
                text_channel = None
                for channel in guild.text_channels:
                    if channel.name == "🔒匿名文字區" and channel.category == category:
                        # 檢查是否是同一個群組的文字頻道（通過檢查權限）
                        text_channel = channel
                        break
                
                if text_channel and non_admin_participants:
                    mentions = " ".join([f"<@{pid}>" for pid in non_admin_participants])
                    await text_channel.send(f"👥 邀請成員：{mentions}")
                
                print(f"✅ 已創建群組語音頻道: {voice_channel.name} ({voice_channel.id})")
                return voice_channel.id
            except Exception as e:
                print(f"❌ 創建群組語音頻道失敗: {e}")
                return None
        
        # 在事件循環中執行異步操作
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        channel_id = loop.run_until_complete(create_channel_async())
        loop.close()
        
        if channel_id:
            return jsonify({
                'success': True,
                'channelId': str(channel_id)
            })
        else:
            return jsonify({'error': '創建頻道失敗'}), 500
            
    except Exception as e:
        print(f"❌ 創建群組語音頻道時發生錯誤: {e}")
        return jsonify({'error': '創建頻道時發生錯誤'}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """健康檢查端點"""
    try:
        # 檢查資料庫連線
        session = Session()
        session.execute(text("SELECT 1"))
        session.close()
        
        return jsonify({
            'status': 'healthy',
            'bot_online': bot.is_ready(),
            'guild_count': len(bot.guilds),
            'tasks_running': {
                'check_bookings': check_bookings.is_running(),
                'check_new_bookings': check_new_bookings.is_running(),
                'database_health_check': database_health_check.is_running()
            }
        })
        except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 500

# 啟動 Flask 應用程式
def run_flask():
    """在單獨的線程中運行 Flask 應用程式"""
    app.run(host='0.0.0.0', port=5001, debug=False, use_reloader=False)

# 主函數
async def main():
    """主函數"""
    print("🚀 啟動 PeiPlay Discord Bot...")
    
    # 在單獨的線程中啟動 Flask
    import threading
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.daemon = True
    flask_thread.start()
    print("✅ Flask API 伺服器已啟動 (端口 5001)")
    
    # 啟動 Discord Bot
    try:
        await bot.start(DISCORD_TOKEN)
        except Exception as e:
        print(f"❌ 啟動 Discord Bot 失敗: {e}")

if __name__ == "__main__":
    asyncio.run(main())